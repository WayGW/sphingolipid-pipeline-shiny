# -*- coding: utf-8 -*-
"""
Sphingolipid Analysis Pipeline - Shiny for Python
===================================================

Run with: shiny run app.py
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from io import BytesIO
from pathlib import Path
import sys
import tempfile
import zipfile
import os
from datetime import datetime

from shiny import App, ui, render, reactive, req

sys.path.insert(0, str(Path(__file__).parent))

from config.sphingolipid_species import (
    SPHINGOLIPID_PANEL, get_ceramides, get_dihydroceramides,
    get_sphingomyelins, get_sphingoid_bases, get_sphingoid_base_phosphates,
    get_saturated, get_unsaturated, get_very_long_chain, get_long_chain
)
from modules.data_processing import SphingolipidDataProcessor, ProcessedData, validate_data_quality
from modules.statistical_tests import StatisticalAnalyzer, format_analysis_report
from modules.visualization import SphingolipidVisualizer, create_summary_figure
from modules.report_generation import (
    ExcelReportGenerator, SignificancePlotter,
    ComprehensiveAnalysisResults, format_apa_statistics,
    get_significant_differences_summary
)


# =============================================================================
# Helper functions (non-UI)
# =============================================================================

def create_excel_with_lod_highlighting(
    df: pd.DataFrame,
    lod_rows: dict,
    sheet_name: str = "Data"
) -> BytesIO:
    """Create an Excel file with cells highlighted where LOD values were replaced."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        from openpyxl.styles import PatternFill, Font
        lod_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        col_positions = {col: idx + 1 for idx, col in enumerate(df.columns)}
        for col_name, row_indices in lod_rows.items():
            if col_name in col_positions:
                col_idx = col_positions[col_name]
                for row_idx in row_indices:
                    excel_row = row_idx + 2
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.fill = lod_fill
        legend_ws = writer.book.create_sheet("Legend")
        legend_ws['A1'] = "Cell Highlighting Legend"
        legend_ws['A1'].font = Font(bold=True, size=12)
        legend_ws['A3'] = "Yellow cells"
        legend_ws['B3'] = "Below LOD - value was replaced based on LOD handling setting"
        legend_ws['A3'].fill = lod_fill
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    buf.seek(0)
    return buf


def create_full_data_excel_with_highlighting(processed, metadata) -> BytesIO:
    """Create a comprehensive Excel file with all data sheets and LOD highlighting."""
    buf = BytesIO()
    lod_rows = processed.structure.analyte_lod_rows
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        from openpyxl.styles import PatternFill, Font
        lod_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        conc_df = pd.concat([metadata, processed.concentrations], axis=1)
        conc_df.to_excel(writer, sheet_name="Concentrations", index=False)
        ws_conc = writer.sheets["Concentrations"]
        conc_col_positions = {col: idx + 1 for idx, col in enumerate(conc_df.columns)}
        for col_name, row_indices in lod_rows.items():
            if col_name in conc_col_positions:
                col_idx = conc_col_positions[col_name]
                for row_idx in row_indices:
                    excel_row = row_idx + 2
                    cell = ws_conc.cell(row=excel_row, column=col_idx)
                    cell.fill = lod_fill
        totals_df = pd.concat([metadata, processed.totals], axis=1)
        totals_df.to_excel(writer, sheet_name="Totals", index=False)
        ratios_df = pd.concat([metadata, processed.ratios], axis=1)
        ratios_df.to_excel(writer, sheet_name="Ratios", index=False)
        pct_df = pd.concat([metadata, processed.percentages], axis=1)
        pct_df.to_excel(writer, sheet_name="Percentages", index=False)
        n_samples = len(processed.sample_data)
        lod_summary = []
        for analyte, lod in sorted(processed.structure.analyte_lods.items()):
            count = processed.structure.analyte_lod_counts.get(analyte, 0)
            pct = (count / n_samples * 100) if n_samples > 0 else 0
            lod_summary.append({
                "Analyte": analyte,
                "LOD (ng/mL)": lod,
                "Below LOD Count": count,
                "% Replaced": round(pct, 1)
            })
        lod_summary_df = pd.DataFrame(lod_summary)
        lod_summary_df.to_excel(writer, sheet_name="LOD Summary", index=False)
        legend_ws = writer.book.create_sheet("Legend")
        legend_ws['A1'] = "Cell Highlighting Legend"
        legend_ws['A1'].font = Font(bold=True, size=12)
        legend_ws['A3'] = "Yellow cells"
        legend_ws['B3'] = "Below LOD - value was replaced based on LOD handling setting"
        legend_ws['A3'].fill = lod_fill
        legend_ws['A5'] = "LOD Source"
        legend_ws['B5'] = processed.structure.lod_source.capitalize()
        legend_ws['A6'] = "Total Samples"
        legend_ws['B6'] = n_samples
        for sn in writer.sheets:
            ws = writer.sheets[sn]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    buf.seek(0)
    return buf


def fig_to_bytes(fig, format='png', dpi=300):
    """Convert matplotlib figure to bytes."""
    buf = BytesIO()
    fig.savefig(buf, format=format, dpi=dpi, bbox_inches='tight')
    buf.seek(0)
    return buf.getvalue()


def get_top_analytes(processed, n=10):
    """Get top N analytes by mean concentration, excluding LOD-saturated species."""
    conc = processed.concentrations
    lod_counts = processed.structure.analyte_lod_counts
    n_samples = len(conc)
    if n_samples == 0:
        return conc.mean().nlargest(n).index.tolist()
    valid_cols = []
    for col in conc.columns:
        n_lod = lod_counts.get(col, 0)
        if n_lod / n_samples <= 0.8:
            valid_cols.append(col)
    if len(valid_cols) >= n:
        return conc[valid_cols].mean().nlargest(n).index.tolist()
    elif valid_cols:
        remaining = [c for c in conc.columns if c not in valid_cols]
        top_valid = conc[valid_cols].mean().nlargest(len(valid_cols)).index.tolist()
        top_remaining = conc[remaining].mean().nlargest(n - len(valid_cols)).index.tolist()
        return top_valid + top_remaining
    else:
        return conc.mean().nlargest(n).index.tolist()


def get_sig_pairs(result, max_pairs=5):
    """Extract significant pairs from result."""
    if not result or not result.posthoc_test or result.posthoc_test.pairwise_results is None:
        return []
    pairs = []
    df = result.posthoc_test.pairwise_results
    sig = df[df['significant']].sort_values('pvalue_adj' if 'pvalue_adj' in df else 'pvalue')
    for _, row in sig.head(max_pairs).iterrows():
        p = row.get('pvalue_adj', row.get('pvalue', 1.0))
        annot = '***' if p < 0.001 else ('**' if p < 0.01 else '*' if p < 0.05 else '')
        if annot:
            pairs.append((row['group1'], row['group2'], annot))
    return pairs


def generate_all_export_figures(processed, results, settings):
    """Generate all figure variations for export package."""
    figures = {}
    viz = SphingolipidVisualizer(color_palette=settings['color_palette'], style=settings['plot_style'])
    group_col = processed.structure.group_col
    if not group_col or not results:
        return figures
    data = processed.sample_data[processed.sample_data[group_col].notna()].copy()
    data = data[data[group_col].astype(str).str.lower() != 'nan']
    available_sls = processed.structure.sphingolipid_cols
    is_twoway = results.is_twoway

    if is_twoway:
        fa_col = results.factor_a_col
        fb_col = results.factor_b_col
        fa_name = results.factor_a_name
        fb_name = results.factor_b_name
        conc_selections = {
            'top10': get_top_analytes(processed, 10),
            'significant': [s for s in available_sls if s in results.twoway_individual_sl
                           and (results.twoway_individual_sl[s].twoway_result.factor_a_pvalue < settings['alpha']
                                or results.twoway_individual_sl[s].twoway_result.factor_b_pvalue < settings['alpha']
                                or results.twoway_individual_sl[s].twoway_result.interaction_pvalue < settings['alpha'])][:10],
            'ceramides': [s for s in get_ceramides() if s in available_sls][:10],
            'sphingomyelins': [s for s in get_sphingomyelins() if s in available_sls][:10],
        }
        for sel_name, selected in conc_selections.items():
            if not selected:
                continue
            tw_stats = {s: results.twoway_individual_sl.get(s) for s in selected
                       if s in results.twoway_individual_sl}
            suffix = f"_{sel_name}"
            try:
                fig = viz.plot_twoway_multi_panel(
                    data=data, value_cols=selected,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (ng/mL)', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures[f'concentrations{suffix}'] = fig
            except Exception:
                pass
            try:
                fig_int = viz.plot_twoway_interaction_multi_panel(
                    data=data, value_cols=selected,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (ng/mL)')
                figures[f'concentrations{suffix}_interaction'] = fig_int
            except Exception:
                pass
        key_totals = ['total_all', 'total_ceramides', 'total_dihydroceramides', 'total_sphingomyelins',
                      'sphingoid_bases', 'sphingoid_base_phosphates', 'very_long_chain', 'long_chain']
        available_totals = [t for t in key_totals if t in processed.totals.columns]
        if available_totals:
            totals_data = pd.concat([data[[fa_col, fb_col]].reset_index(drop=True),
                                     processed.totals[available_totals].loc[data.index].reset_index(drop=True)], axis=1)
            tw_stats = {t: results.twoway_totals.get(t) for t in available_totals
                       if t in results.twoway_totals}
            try:
                fig = viz.plot_twoway_multi_panel(
                    data=totals_data, value_cols=available_totals,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (ng/mL)', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures['totals'] = fig
            except Exception:
                pass
            try:
                fig_int = viz.plot_twoway_interaction_multi_panel(
                    data=totals_data, value_cols=available_totals,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Concentration (ng/mL)')
                figures['totals_interaction'] = fig_int
            except Exception:
                pass
        pct_cols = [col for col in processed.percentages.columns if col.endswith('_pct')]
        if pct_cols:
            top_pct = processed.percentages[pct_cols].mean().nlargest(10).index.tolist()
            display_map = {col: col.replace('_pct', '') for col in top_pct}
            display_cols = list(display_map.values())
            pct_data = pd.concat([data[[fa_col, fb_col]].reset_index(drop=True),
                                  processed.percentages[top_pct].loc[data.index].reset_index(drop=True)], axis=1)
            pct_data = pct_data.rename(columns=display_map)
            tw_stats = {}
            for pct_col in top_pct:
                disp = pct_col.replace('_pct', '')
                if pct_col in results.twoway_percentages:
                    tw_stats[disp] = results.twoway_percentages[pct_col]
            try:
                fig = viz.plot_twoway_multi_panel(
                    data=pct_data, value_cols=display_cols,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='% of Total SL', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures['percentages_top10'] = fig
            except Exception:
                pass
            try:
                fig_int = viz.plot_twoway_interaction_multi_panel(
                    data=pct_data, value_cols=display_cols,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='% of Total SL')
                figures['percentages_interaction'] = fig_int
            except Exception:
                pass
        ratio_cols = [col for col in processed.ratios.columns if not processed.ratios[col].isna().all()]
        if ratio_cols:
            selected_ratios = ratio_cols[:9]
            ratio_data = pd.concat([data[[fa_col, fb_col]].reset_index(drop=True),
                                    processed.ratios[selected_ratios].loc[data.index].reset_index(drop=True)], axis=1)
            tw_stats = {r: results.twoway_ratios.get(r) for r in selected_ratios
                       if r in results.twoway_ratios}
            try:
                fig = viz.plot_twoway_multi_panel(
                    data=ratio_data, value_cols=selected_ratios,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Ratio', show_points=settings['show_points'],
                    plot_type=settings['plot_type'])
                figures['ratios'] = fig
            except Exception:
                pass
            try:
                fig_int = viz.plot_twoway_interaction_multi_panel(
                    data=ratio_data, value_cols=selected_ratios,
                    factor_a_col=fa_col, factor_b_col=fb_col,
                    twoway_results=tw_stats, ncols=3,
                    factor_a_name=fa_name, factor_b_name=fb_name,
                    ylabel='Ratio')
                figures['ratios_interaction'] = fig_int
            except Exception:
                pass
    else:
        conc_selections = {
            'top10': get_top_analytes(processed, 10),
            'significant': [s for s in available_sls if s in results.individual_sl_results
                           and results.individual_sl_results[s].main_test.significant][:10],
            'ceramides': [s for s in get_ceramides() if s in available_sls][:10],
            'sphingomyelins': [s for s in get_sphingomyelins() if s in available_sls][:10],
        }
        for sel_name, selected in conc_selections.items():
            if not selected:
                continue
            stats_dict = {s: results.individual_sl_results.get(s) for s in selected if s in results.individual_sl_results}
            for log_scale in [False, True]:
                suffix = f"_{sel_name}{'_log' if log_scale else ''}"
                try:
                    fig = viz.plot_multi_panel_groups_with_stats(
                        data, selected, group_col, stats_dict, ncols=3,
                        plot_type=settings['plot_type'], log_scale=log_scale,
                        show_points=settings['show_points'])
                    figures[f'concentrations{suffix}'] = fig
                except Exception:
                    pass
        totals_combined = pd.concat([data[[group_col]], processed.totals.loc[data.index]], axis=1)
        key_totals = ['total_all', 'total_ceramides', 'total_dihydroceramides', 'total_sphingomyelins',
                      'sphingoid_bases', 'sphingoid_base_phosphates', 'very_long_chain', 'long_chain']
        available_totals = [t for t in key_totals if t in totals_combined.columns]
        if available_totals:
            totals_stats = {t: results.totals_results.get(t) for t in available_totals if t in results.totals_results}
            for log_scale in [False, True]:
                suffix = '_log' if log_scale else ''
                try:
                    fig = viz.plot_multi_panel_groups_with_stats(
                        totals_combined, available_totals, group_col, totals_stats, ncols=3,
                        plot_type=settings['plot_type'], log_scale=log_scale,
                        show_points=settings['show_points'])
                    figures[f'totals{suffix}'] = fig
                except Exception:
                    pass
        pct_cols = [col for col in processed.percentages.columns if col.endswith('_pct')]
        if pct_cols:
            top_pct = processed.percentages[pct_cols].mean().nlargest(10).index.tolist()
            pct_combined = pd.concat([data[[group_col]], processed.percentages.loc[data.index, top_pct]], axis=1)
            display_cols = [col.replace('_pct', '') for col in top_pct]
            pct_display = pct_combined.rename(columns={old: new for old, new in zip(top_pct, display_cols)})
            pct_stats = {}
            for pct_col, disp_col in zip(top_pct, display_cols):
                if pct_col in results.percentages_results:
                    pct_stats[disp_col] = results.percentages_results[pct_col]
            try:
                fig = viz.plot_multi_panel_groups_with_stats(
                    pct_display, display_cols, group_col, pct_stats, ncols=3,
                    plot_type=settings['plot_type'], log_scale=False,
                    show_points=settings['show_points'])
                figures['percentages_top10'] = fig
            except Exception:
                pass
        ratio_cols = [col for col in processed.ratios.columns if not processed.ratios[col].isna().all()]
        if ratio_cols:
            ratios_combined = pd.concat([data[[group_col]], processed.ratios.loc[data.index, ratio_cols]], axis=1)
            ratios_stats = {r: results.ratios_results.get(r) for r in ratio_cols if r in results.ratios_results}
            for log_scale in [False, True]:
                suffix = '_log' if log_scale else ''
                try:
                    fig = viz.plot_multi_panel_groups_with_stats(
                        ratios_combined, ratio_cols[:9], group_col, ratios_stats, ncols=3,
                        plot_type=settings['plot_type'], log_scale=log_scale,
                        show_points=settings['show_points'])
                    figures[f'ratios{suffix}'] = fig
                except Exception:
                    pass
    return figures


def create_results_zip(processed, results, figures, report_gen, metadata):
    """Create ZIP with all results including LOD-highlighted Excel files."""
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        full_excel_buf = create_full_data_excel_with_highlighting(processed, metadata)
        zf.writestr('data/sphingolipid_all_data.xlsx', full_excel_buf.getvalue())
        for name, df in [('concentrations', processed.concentrations),
                         ('percentages', processed.percentages),
                         ('totals', processed.totals),
                         ('ratios', processed.ratios)]:
            csv_buf = BytesIO()
            export_df = pd.concat([metadata, df], axis=1)
            export_df.to_csv(csv_buf, index=False)
            zf.writestr(f'data/{name}.csv', csv_buf.getvalue())
        full = pd.concat([processed.sample_data, processed.totals, processed.ratios, processed.percentages], axis=1)
        csv_buf = BytesIO()
        full.to_csv(csv_buf, index=False)
        zf.writestr('data/full_analysis.csv', csv_buf.getvalue())
        if report_gen:
            excel_buf = BytesIO()
            report_gen.save_excel_report(excel_buf)
            excel_buf.seek(0)
            zf.writestr('reports/statistical_report.xlsx', excel_buf.getvalue())
        for name, fig in figures.items():
            if fig:
                zf.writestr(f'figures/{name}.pdf', fig_to_bytes(fig, 'pdf'))
        readme_content = f"""Sphingolipid Analysis Results
==============================

Generated: {datetime.now():%Y-%m-%d %H:%M}

Contents:
---------
data/
  - sphingolipid_all_data.xlsx  (Primary data file with LOD highlighting)
  - concentrations.csv
  - totals.csv
  - ratios.csv
  - percentages.csv
  - full_analysis.csv

reports/
  - statistical_report.xlsx

figures/
  - Various PDF figures

LOD Highlighting:
-----------------
In sphingolipid_all_data.xlsx, yellow cells indicate values that were
below the limit of detection (LOD) and were replaced according to the
selected LOD handling method.

LOD Source: {processed.structure.lod_source}
Total Samples: {len(processed.sample_data)}
Total Analytes: {len(processed.structure.sphingolipid_cols)}
"""
        zf.writestr('README.txt', readme_content)
    buf.seek(0)
    return buf.getvalue()


def _get_metadata(processed):
    """Get metadata columns (Sample_ID and Group)."""
    id_cols = []
    if processed.structure.sample_id_col and processed.structure.sample_id_col in processed.sample_data.columns:
        id_cols.append(processed.structure.sample_id_col)
    if processed.structure.group_col and processed.structure.group_col in processed.sample_data.columns:
        id_cols.append(processed.structure.group_col)
    return processed.sample_data[id_cols] if id_cols else pd.DataFrame(index=processed.sample_data.index)


# =============================================================================
# UI Definition
# =============================================================================

conc_tab = ui.nav_panel(
    "Concentrations",
    ui.h3("Individual Sphingolipid Concentrations"),
    ui.layout_columns(
        ui.input_select("conc_quick", "Quick select",
                        ["Top 10", "Significant", "Ceramides", "Sphingomyelins", "Custom"]),
        ui.panel_conditional(
            "input.conc_quick === 'Custom'",
            ui.output_ui("conc_custom_select"),
        ),
        col_widths=[4, 8],
    ),
    ui.input_checkbox("conc_log", "Log10 scale"),
    ui.output_ui("conc_twoway_label"),
    ui.output_plot("conc_plot", height="700px"),
    ui.accordion(
        ui.accordion_panel("Interaction Plots", ui.output_plot("conc_interaction_plot", height="600px")),
        ui.accordion_panel("Statistical Summary", ui.output_data_frame("conc_stats_table")),
        open=False,
        id="conc_accordion",
    ),
)

totals_tab = ui.nav_panel(
    "Totals",
    ui.h3("Aggregate Totals"),
    ui.input_checkbox("totals_log", "Log10 scale"),
    ui.output_plot("totals_plot", height="700px"),
    ui.accordion(
        ui.accordion_panel("Interaction Plots", ui.output_plot("totals_interaction_plot", height="600px")),
        ui.accordion_panel("Statistical Summary", ui.output_data_frame("totals_stats_table")),
        open=False,
        id="totals_accordion",
    ),
)

pct_tab = ui.nav_panel(
    "Percentages",
    ui.h3("Sphingolipid Pool Composition (%)"),
    ui.p("Compare the percentage each sphingolipid contributes to the total pool across groups",
         style="color: #666; font-size: 0.9em;"),
    ui.layout_columns(
        ui.input_select("pct_quick", "Quick select",
                        ["Top 10 by mean %", "Significant", "Ceramides", "Dihydroceramides",
                         "Sphingomyelins", "Very Long Chain", "Custom"]),
        ui.panel_conditional(
            "input.pct_quick === 'Custom'",
            ui.output_ui("pct_custom_select"),
        ),
        col_widths=[4, 8],
    ),
    ui.output_ui("pct_sig_info"),
    ui.h4("Statistical Comparison by Group"),
    ui.output_plot("pct_plot", height="700px"),
    ui.accordion(
        ui.accordion_panel("Interaction Plots", ui.output_plot("pct_interaction_plot", height="600px")),
        open=False,
        id="pct_interaction_accordion",
    ),
    ui.hr(),
    ui.h4("Pool Composition - Pie Charts"),
    ui.output_plot("pct_pie_plot", height="500px"),
    ui.hr(),
    ui.h4("Pool Composition - Bar Comparison"),
    ui.output_plot("pct_bars_plot", height="500px"),
    ui.hr(),
    ui.h4("Stacked Composition View"),
    ui.output_plot("pct_stacked_plot", height="400px"),
    ui.accordion(
        ui.accordion_panel("Statistical Summary", ui.output_data_frame("pct_stats_table")),
        ui.accordion_panel("Group Mean Percentages", ui.output_data_frame("pct_means_table")),
        ui.accordion_panel("View percentage data", ui.output_data_frame("pct_raw_table")),
        open=False,
        id="pct_accordion",
    ),
)

ratios_tab = ui.nav_panel(
    "Ratios",
    ui.h3("Clinical Ratios"),
    ui.layout_columns(
        ui.input_select("ratio_quick", "Quick select",
                        ["All ratios", "Significant only", "Key ratios", "Custom"]),
        ui.panel_conditional(
            "input.ratio_quick === 'Custom'",
            ui.output_ui("ratio_custom_select"),
        ),
        col_widths=[4, 8],
    ),
    ui.output_ui("ratio_sig_info"),
    ui.input_checkbox("ratio_log", "Log10 scale"),
    ui.output_plot("ratio_plot", height="700px"),
    ui.accordion(
        ui.accordion_panel("Interaction Plots", ui.output_plot("ratio_interaction_plot", height="600px")),
        ui.accordion_panel("Statistical Summary", ui.output_data_frame("ratio_stats_table")),
        ui.accordion_panel("Group Mean Ratios", ui.output_data_frame("ratio_means_table")),
        ui.accordion_panel("View ratio data", ui.output_data_frame("ratio_raw_table")),
        open=False,
        id="ratio_accordion",
    ),
)

stats_tab = ui.nav_panel(
    "Statistics",
    ui.h3("Statistics Summary"),
    ui.output_ui("stats_content"),
)

export_tab = ui.nav_panel(
    "Export",
    ui.h3("Export Results"),
    ui.layout_columns(
        ui.card(
            ui.card_header("Individual Downloads"),
            ui.download_button("download_concentrations", "Concentrations (Excel)"),
            ui.download_button("download_totals", "Totals (CSV)"),
            ui.download_button("download_ratios", "Ratios (CSV)"),
            ui.download_button("download_percentages", "Percentages (CSV)"),
        ),
        ui.card(
            ui.card_header("Complete Reports"),
            ui.download_button("download_full_excel", "All Data (Excel)"),
            ui.download_button("download_stats_report", "Statistical Report (Excel)"),
            ui.download_button("download_zip", "Complete Package (ZIP)"),
        ),
        col_widths=[6, 6],
    ),
    ui.hr(),
    ui.card(
        ui.card_header("Summary Figure"),
        ui.input_action_button("gen_summary", "Generate Summary Figure"),
        ui.output_plot("summary_plot", height="600px"),
        ui.layout_columns(
            ui.download_button("download_summary_png", "Summary PNG"),
            ui.download_button("download_summary_pdf", "Summary PDF"),
            col_widths=[6, 6],
        ),
    ),
)

app_ui = ui.page_sidebar(
    ui.sidebar(
        ui.h4("Settings"),
        ui.hr(),
        ui.h5("Detection Limits"),
        ui.p("LODs are auto-detected per-analyte from standard curves", style="color: #666; font-size: 0.85em;"),
        ui.input_numeric("lod_value", "Fallback LOD (ng/mL)", value=1.0, min=0.0, max=100.0, step=0.5),
        ui.input_select("lod_handling", "Below LOD handling",
                        {"lod": "LOD value", "half_lod": "LOD/2", "zero": "Zero",
                         "half_min": "Min/2", "drop": "NaN"}),
        ui.hr(),
        ui.h5("Analysis"),
        ui.input_slider("alpha", "Significance (alpha)", min=0.01, max=0.10, value=0.05, step=0.01),
        ui.input_select("plot_type", "Plot type",
                        {"box": "Box", "violin": "Violin", "bar": "Bar", "strip": "Strip"}),
        ui.input_checkbox("show_points", "Show data points", value=True),
        ui.hr(),
        ui.h5("Appearance"),
        ui.input_select("color_palette", "Color palette",
                        {"Set2": "Set2 (Default - Soft pastels)", "Set1": "Set1 (Bold primary)",
                         "Paired": "Paired (Light/dark pairs)", "Dark2": "Dark2 (Darker pastels)",
                         "colorblind": "Colorblind-friendly", "tab10": "Tab10 (Matplotlib default)",
                         "Pastel1": "Pastel1 (Very soft)", "Accent": "Accent (High contrast)"}),
        ui.input_select("plot_style", "Plot background",
                        {"whitegrid": "White with grid", "white": "Clean white",
                         "darkgrid": "Gray with grid", "ticks": "White with ticks"}),
        width=280,
    ),
    ui.h2("Sphingolipid Analysis Pipeline"),
    ui.input_file("file", "Upload Excel/ODS file", accept=[".xlsx", ".xls", ".ods"]),
    ui.output_ui("sheet_selector"),
    ui.output_ui("data_summary"),
    ui.navset_tab(
        conc_tab,
        totals_tab,
        pct_tab,
        ratios_tab,
        stats_tab,
        export_tab,
        id="main_tabs",
    ),
    ui.output_ui("upload_instructions"),
    title="Sphingolipid Analysis Pipeline",
)


# =============================================================================
# Server
# =============================================================================

def server(input, output, session):

    # =========================================================================
    # Core reactive calculations
    # =========================================================================

    @reactive.calc
    def settings():
        return {
            'lod_handling': input.lod_handling(),
            'lod_value': input.lod_value(),
            'alpha': input.alpha(),
            'plot_type': input.plot_type(),
            'show_points': input.show_points(),
            'color_palette': input.color_palette(),
            'plot_style': input.plot_style(),
        }

    @reactive.calc
    def file_info():
        f = input.file()
        req(f)
        return f[0]

    @reactive.calc
    def available_sheets():
        fi = file_info()
        processor = SphingolipidDataProcessor(lod_handling="lod", lod_value=1.0)
        sheets, auto = processor.get_selectable_sheets(fi["datapath"])
        return sheets, auto

    @reactive.calc
    def processed_data():
        fi = file_info()
        sheets, auto = available_sheets()
        if len(sheets) > 1:
            sheet = input.sheet()
            if not sheet:
                sheet = auto
        else:
            sheet = sheets[0] if sheets else None
        req(sheet)
        s = settings()
        processor = SphingolipidDataProcessor(
            lod_handling=s['lod_handling'],
            lod_value=s['lod_value'],
        )
        return processor.load_and_process(fi["datapath"], sheet_name=sheet)

    @reactive.calc
    def analysis_results():
        """Compute all statistics (memoized by reactive dependencies)."""
        proc = processed_data()
        req(proc)
        group_col = proc.structure.group_col
        req(group_col)
        s = settings()
        factors = getattr(proc.structure, 'factors', {})
        n_factors = getattr(proc.structure, 'n_factors', 0)
        report_gen = ExcelReportGenerator(
            data=proc.sample_data,
            group_col=group_col,
            sphingolipid_cols=proc.structure.sphingolipid_cols,
            totals=proc.totals,
            ratios=proc.ratios,
            percentages=proc.percentages,
            alpha=s['alpha'],
            factors=factors,
            n_factors=n_factors,
        )
        results = report_gen.run_all_statistics()
        return results, report_gen

    @reactive.calc
    def results_obj():
        r = analysis_results()
        req(r)
        return r[0]

    @reactive.calc
    def report_gen():
        r = analysis_results()
        req(r)
        return r[1]

    @reactive.calc
    def is_twoway():
        r = results_obj()
        return r is not None and r.is_twoway

    @reactive.calc
    def clean_data():
        """Get filtered sample data (no NaN groups)."""
        proc = processed_data()
        req(proc)
        group_col = proc.structure.group_col
        req(group_col)
        data = proc.sample_data[proc.sample_data[group_col].notna()].copy()
        data = data[data[group_col].astype(str).str.lower() != 'nan']
        return data

    def _make_viz():
        s = settings()
        return SphingolipidVisualizer(color_palette=s['color_palette'], style=s['plot_style'])

    # =========================================================================
    # Dynamic UI
    # =========================================================================

    @render.ui
    def sheet_selector():
        try:
            sheets, auto = available_sheets()
        except Exception:
            return ui.TagList()
        if len(sheets) > 1:
            return ui.input_select("sheet", "Select data sheet", choices=sheets, selected=auto)
        return ui.TagList()

    @render.ui
    def data_summary():
        try:
            proc = processed_data()
        except Exception:
            return ui.TagList()
        req(proc)
        quality = validate_data_quality(proc)
        s = settings()
        lod_display = {"lod": "LOD value", "half_lod": "LOD/2", "zero": "Zero",
                       "half_min": "Min/2", "drop": "NaN"}
        lod_source = proc.structure.lod_source
        n_lods = len(proc.structure.analyte_lods)
        n_factors = getattr(proc.structure, 'n_factors', 0)

        elements = [
            ui.layout_columns(
                ui.value_box("Samples", quality['n_samples']),
                ui.value_box("Sphingolipids", quality['n_sphingolipids_detected']),
                ui.value_box("Groups", quality.get('n_groups', 'N/A')),
                col_widths=[4, 4, 4],
            ),
        ]

        if lod_source == "standards":
            elements.append(ui.p(
                f"LODs: Auto-detected from standard curves ({n_lods} analytes) | "
                f"Below LOD -> {lod_display[s['lod_handling']]} | alpha = {s['alpha']}",
                style="color: #666; font-size: 0.85em;"
            ))
        else:
            elements.append(ui.p(
                f"LODs: Default ({s['lod_value']} ng/mL) | "
                f"Below LOD -> {lod_display[s['lod_handling']]} | alpha = {s['alpha']}",
                style="color: #666; font-size: 0.85em;"
            ))

        if n_factors >= 2:
            factors = proc.structure.factors
            factor_names = list(factors.keys())
            factor_source = getattr(proc.structure, 'factor_source', 'unknown')
            elements.append(ui.div(
                ui.strong(f"Two-Way Design Detected: {factor_names[0]} x {factor_names[1]} "
                          f"(from {factor_source.replace('_', ' ')})"),
                class_="alert alert-info",
            ))

        # LOD details accordion
        if proc.structure.analyte_lods:
            n_samples = quality['n_samples']
            lod_counts = proc.structure.analyte_lod_counts
            lod_data = []
            for analyte, lod in sorted(proc.structure.analyte_lods.items()):
                count = lod_counts.get(analyte, 0)
                pct = (count / n_samples * 100) if n_samples > 0 else 0
                lod_data.append({
                    "Analyte": analyte, "LOD (ng/mL)": lod,
                    "Below LOD": count, "% Replaced": f"{pct:.1f}%"
                })
            high_replacement = [d["Analyte"] for d in lod_data if float(d["% Replaced"].rstrip('%')) > 50]
            warning_el = []
            if high_replacement:
                warning_el = [ui.div(
                    ui.strong(f"High replacement rate (>50%): "
                              f"{', '.join(high_replacement[:5])}{'...' if len(high_replacement) > 5 else ''}"),
                    class_="alert alert-warning",
                )]
            elements.append(
                ui.accordion(
                    ui.accordion_panel(
                        "View Per-Analyte LODs",
                        ui.output_data_frame("lod_detail_table"),
                        *warning_el,
                    ),
                    open=False,
                    id="lod_accordion",
                )
            )

        return ui.TagList(*elements)

    @render.data_frame
    def lod_detail_table():
        proc = processed_data()
        req(proc, proc.structure.analyte_lods)
        quality = validate_data_quality(proc)
        n_samples = quality['n_samples']
        lod_counts = proc.structure.analyte_lod_counts
        rows = []
        for analyte, lod in sorted(proc.structure.analyte_lods.items()):
            count = lod_counts.get(analyte, 0)
            pct = (count / n_samples * 100) if n_samples > 0 else 0
            rows.append({"Analyte": analyte, "LOD (ng/mL)": lod,
                         "Below LOD": count, "% Replaced": f"{pct:.1f}%"})
        return render.DataGrid(pd.DataFrame(rows))

    @render.ui
    def upload_instructions():
        f = input.file()
        if f:
            return ui.TagList()
        example_df = pd.DataFrame({
            'Type': ['Aged', 'Aged', 'Young'],
            'Sample_ID': ['S001', 'S002', 'S003'],
            'C16 Cer': ['125.4', '142.8', '98.6'],
            'C24-0 Cer': ['312.5', '287.9', '-----'],
            'C16-SM': ['1520.3', '1380.7', '1245.2'],
            'S-d18-1': ['15.2', '18.4', '12.8'],
        })
        example_tw = pd.DataFrame({
            'Factor_Age': ['Aged', 'Aged', 'Young'],
            'Factor_Sex': ['Male', 'Female', 'Female'],
            'Sample_ID': ['S001', 'S002', 'S003'],
            'C16 Cer': ['125.4', '142.8', '98.6'],
            'C24-0 Cer': ['312.5', '287.9', '-----'],
            'C16-SM': ['1520.3', '1380.7', '1245.2'],
        })
        return ui.TagList(
            ui.hr(),
            ui.h3("Expected Input Format"),
            ui.h4("Sample sheet"),
            ui.tags.ul(
                ui.tags.li("Rows: Samples"),
                ui.tags.li("Columns: Sphingolipid species (matching panel names)"),
                ui.tags.li("First column(s): Sample ID, Group/Type"),
                ui.tags.li("Values: Concentrations (typically ng/mL)"),
                ui.tags.li('Below LOD: Can be "-----", "LOD", "BLQ", "ND", etc.'),
            ),
            ui.p(ui.strong('For Multiple Independent Variables add "Factor_" in front of it')),
            ui.h4("Example:"),
            ui.output_data_frame("example_table_1"),
            ui.h4("Two-factor example:"),
            ui.output_data_frame("example_table_2"),
            ui.div(
                ui.tags.ul(
                    ui.tags.li("Sphingolipid columns are auto-detected from your data"),
                    ui.tags.li("Group assignments are inferred automatically"),
                    ui.tags.li(ui.strong("Per-analyte LODs"), " are extracted from standard curves in the 'LC-MS data' sheet"),
                ),
                class_="alert alert-info",
            ),
        )

    @render.data_frame
    def example_table_1():
        return render.DataGrid(pd.DataFrame({
            'Type': ['Aged', 'Aged', 'Young'],
            'Sample_ID': ['S001', 'S002', 'S003'],
            'C16 Cer': ['125.4', '142.8', '98.6'],
            'C24-0 Cer': ['312.5', '287.9', '-----'],
            'C16-SM': ['1520.3', '1380.7', '1245.2'],
        }))

    @render.data_frame
    def example_table_2():
        return render.DataGrid(pd.DataFrame({
            'Factor_Age': ['Aged', 'Aged', 'Young'],
            'Factor_Sex': ['Male', 'Female', 'Female'],
            'Sample_ID': ['S001', 'S002', 'S003'],
            'C16 Cer': ['125.4', '142.8', '98.6'],
            'C24-0 Cer': ['312.5', '287.9', '-----'],
        }))

    # =========================================================================
    # Concentrations tab
    # =========================================================================

    @render.ui
    def conc_custom_select():
        proc = processed_data()
        req(proc)
        return ui.input_selectize("conc_custom", "Select sphingolipids",
                                  choices=list(proc.structure.sphingolipid_cols),
                                  selected=list(proc.structure.sphingolipid_cols[:5]),
                                  multiple=True)

    @reactive.calc
    def selected_concentrations():
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        s = settings()
        quick = input.conc_quick()
        available_sls = proc.structure.sphingolipid_cols

        if quick == "Top 10":
            return get_top_analytes(proc, 10)
        elif quick == "Significant":
            if is_twoway():
                sel = [sl for sl in available_sls if sl in results.twoway_individual_sl
                       and (results.twoway_individual_sl[sl].twoway_result.factor_a_pvalue < s['alpha']
                            or results.twoway_individual_sl[sl].twoway_result.factor_b_pvalue < s['alpha']
                            or results.twoway_individual_sl[sl].twoway_result.interaction_pvalue < s['alpha'])][:10]
            else:
                sel = [sl for sl in available_sls if sl in results.individual_sl_results
                       and results.individual_sl_results[sl].main_test.significant][:10]
            return sel if sel else get_top_analytes(proc, 5)
        elif quick == "Ceramides":
            return [s for s in get_ceramides() if s in available_sls][:10]
        elif quick == "Sphingomyelins":
            return [s for s in get_sphingomyelins() if s in available_sls][:10]
        else:  # Custom
            custom = input.conc_custom()
            return list(custom) if custom else []

    @render.ui
    def conc_twoway_label():
        if not is_twoway():
            return ui.TagList()
        results = results_obj()
        return ui.p(ui.strong(f"Two-way ANOVA: {results.factor_a_name} x {results.factor_b_name}"))

    @render.plot
    def conc_plot():
        proc = processed_data()
        results = results_obj()
        selected = selected_concentrations()
        req(proc, results, selected)
        s = settings()
        data = clean_data()
        viz = _make_viz()
        group_col = proc.structure.group_col

        if is_twoway():
            tw_stats = {sl: results.twoway_individual_sl.get(sl) for sl in selected
                       if sl in results.twoway_individual_sl}
            return viz.plot_twoway_multi_panel(
                data=data, value_cols=selected,
                factor_a_col=results.factor_a_col, factor_b_col=results.factor_b_col,
                twoway_results=tw_stats, ncols=3,
                factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
                ylabel='Concentration (ng/mL)', show_points=s['show_points'],
                plot_type=s['plot_type'])
        else:
            stats_dict = {sl: results.individual_sl_results.get(sl) for sl in selected
                         if sl in results.individual_sl_results}
            return viz.plot_multi_panel_groups_with_stats(
                data, selected, group_col, stats_dict, ncols=3,
                plot_type=s['plot_type'], log_scale=input.conc_log(),
                show_points=s['show_points'], ylabel='Concentration (ng/mL)')

    @render.plot
    def conc_interaction_plot():
        req(is_twoway())
        proc = processed_data()
        results = results_obj()
        selected = selected_concentrations()
        req(proc, results, selected)
        data = clean_data()
        viz = _make_viz()
        tw_stats = {sl: results.twoway_individual_sl.get(sl) for sl in selected
                   if sl in results.twoway_individual_sl}
        return viz.plot_twoway_interaction_multi_panel(
            data=data, value_cols=selected,
            factor_a_col=results.factor_a_col, factor_b_col=results.factor_b_col,
            twoway_results=tw_stats, ncols=3,
            factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
            ylabel='Concentration (ng/mL)')

    @render.data_frame
    def conc_stats_table():
        proc = processed_data()
        results = results_obj()
        selected = selected_concentrations()
        req(proc, results, selected)
        if is_twoway():
            from modules.report_generation import get_twoway_differences_summary
            tw_stats = {sl: results.twoway_individual_sl.get(sl) for sl in selected
                       if sl in results.twoway_individual_sl}
            if tw_stats:
                return render.DataGrid(get_twoway_differences_summary(tw_stats))
            return render.DataGrid(pd.DataFrame())
        else:
            stats_dict = {sl: results.individual_sl_results.get(sl) for sl in selected
                         if sl in results.individual_sl_results}
            return render.DataGrid(get_significant_differences_summary(stats_dict))

    # =========================================================================
    # Totals tab
    # =========================================================================

    @render.plot
    def totals_plot():
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        s = settings()
        data = clean_data()
        viz = _make_viz()
        group_col = proc.structure.group_col
        key_totals = ['total_all', 'total_ceramides', 'total_dihydroceramides', 'total_sphingomyelins',
                      'sphingoid_bases', 'sphingoid_base_phosphates', 'very_long_chain', 'long_chain']

        if is_twoway():
            combined = pd.concat([data[[results.factor_a_col, results.factor_b_col]],
                                  proc.totals.loc[data.index]], axis=1)
            available = [t for t in key_totals if t in combined.columns]
            req(available)
            tw_stats = {t: results.twoway_totals.get(t) for t in available if t in results.twoway_totals}
            return viz.plot_twoway_multi_panel(
                data=combined, value_cols=available,
                factor_a_col=results.factor_a_col, factor_b_col=results.factor_b_col,
                twoway_results=tw_stats, ncols=3,
                factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
                ylabel='Concentration (ng/mL)', show_points=s['show_points'],
                plot_type=s['plot_type'])
        else:
            combined = pd.concat([data[[group_col]], proc.totals.loc[data.index]], axis=1)
            available = [t for t in key_totals if t in combined.columns]
            req(available)
            stats_dict = {t: results.totals_results.get(t) for t in available if t in results.totals_results}
            return viz.plot_multi_panel_groups_with_stats(
                combined, available, group_col, stats_dict, ncols=3,
                plot_type=s['plot_type'], log_scale=input.totals_log(),
                show_points=s['show_points'], ylabel='Concentration (ng/mL)')

    @render.plot
    def totals_interaction_plot():
        req(is_twoway())
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        data = clean_data()
        viz = _make_viz()
        key_totals = ['total_all', 'total_ceramides', 'total_dihydroceramides', 'total_sphingomyelins',
                      'sphingoid_bases', 'sphingoid_base_phosphates', 'very_long_chain', 'long_chain']
        combined = pd.concat([data[[results.factor_a_col, results.factor_b_col]],
                              proc.totals.loc[data.index]], axis=1)
        available = [t for t in key_totals if t in combined.columns]
        req(available)
        tw_stats = {t: results.twoway_totals.get(t) for t in available if t in results.twoway_totals}
        return viz.plot_twoway_interaction_multi_panel(
            data=combined, value_cols=available,
            factor_a_col=results.factor_a_col, factor_b_col=results.factor_b_col,
            twoway_results=tw_stats, ncols=3,
            factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
            ylabel='Concentration (ng/mL)')

    @render.data_frame
    def totals_stats_table():
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        if is_twoway():
            from modules.report_generation import get_twoway_differences_summary
            key_totals = ['total_all', 'total_ceramides', 'total_dihydroceramides', 'total_sphingomyelins',
                          'sphingoid_bases', 'sphingoid_base_phosphates', 'very_long_chain', 'long_chain']
            available = [t for t in key_totals if t in proc.totals.columns]
            tw_stats = {t: results.twoway_totals.get(t) for t in available if t in results.twoway_totals}
            if tw_stats:
                return render.DataGrid(get_twoway_differences_summary(tw_stats))
            return render.DataGrid(pd.DataFrame())
        else:
            rows = []
            key_totals = ['total_all', 'total_ceramides', 'total_dihydroceramides', 'total_sphingomyelins',
                          'sphingoid_bases', 'sphingoid_base_phosphates', 'very_long_chain', 'long_chain']
            available = [t for t in key_totals if t in proc.totals.columns]
            for t in available:
                res = results.totals_results.get(t)
                if res:
                    rows.append({
                        'Total': t.replace('_', ' ').title(),
                        'P-value': f"{res.main_test.pvalue:.4f}",
                        'Significant': 'Yes' if res.main_test.significant else '',
                        'APA': format_apa_statistics(res)
                    })
            return render.DataGrid(pd.DataFrame(rows))

    # =========================================================================
    # Percentages tab
    # =========================================================================

    @render.ui
    def pct_custom_select():
        proc = processed_data()
        req(proc)
        pct_cols = [c for c in proc.percentages.columns if c.endswith('_pct')]
        sl_names = [c.replace('_pct', '') for c in pct_cols]
        return ui.input_selectize("pct_custom", "Select sphingolipids",
                                  choices=sl_names, selected=sl_names[:5], multiple=True)

    @reactive.calc
    def selected_pct_cols():
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        s = settings()
        pct_cols = [c for c in proc.percentages.columns if c.endswith('_pct')]
        quick = input.pct_quick()

        if quick == "Top 10 by mean %":
            return proc.percentages[pct_cols].mean().nlargest(10).index.tolist()
        elif quick == "Significant":
            if is_twoway():
                sel = [c for c in pct_cols if c in results.twoway_percentages
                       and (results.twoway_percentages[c].twoway_result.factor_a_pvalue < s['alpha']
                            or results.twoway_percentages[c].twoway_result.factor_b_pvalue < s['alpha']
                            or results.twoway_percentages[c].twoway_result.interaction_pvalue < s['alpha'])][:10]
            else:
                sel = [c for c in pct_cols if c in results.percentages_results
                       and results.percentages_results[c].main_test.significant][:10]
            if not sel:
                sel = proc.percentages[pct_cols].mean().nlargest(10).index.tolist()
            return sel
        elif quick == "Ceramides":
            return [f'{sl}_pct' for sl in get_ceramides() if f'{sl}_pct' in pct_cols][:10]
        elif quick == "Dihydroceramides":
            return [f'{sl}_pct' for sl in get_dihydroceramides() if f'{sl}_pct' in pct_cols][:10]
        elif quick == "Sphingomyelins":
            return [f'{sl}_pct' for sl in get_sphingomyelins() if f'{sl}_pct' in pct_cols][:10]
        elif quick == "Very Long Chain":
            return [f'{sl}_pct' for sl in get_very_long_chain() if f'{sl}_pct' in pct_cols][:10]
        else:  # Custom
            custom = input.pct_custom()
            return [f'{sl}_pct' for sl in custom] if custom else []

    @render.ui
    def pct_sig_info():
        results = results_obj()
        sel = selected_pct_cols()
        req(results, sel)
        s = settings()
        if is_twoway():
            sig = [c for c in sel if c in results.twoway_percentages
                   and (results.twoway_percentages[c].twoway_result.factor_a_pvalue < s['alpha']
                        or results.twoway_percentages[c].twoway_result.factor_b_pvalue < s['alpha']
                        or results.twoway_percentages[c].twoway_result.interaction_pvalue < s['alpha'])]
        else:
            sig = [c for c in sel if c in results.percentages_results
                   and results.percentages_results[c].main_test.significant]
        if sig:
            sig_names = [c.replace('_pct', '') for c in sig]
            return ui.div(ui.strong(f"Significant differences: {', '.join(sig_names)}"),
                          class_="alert alert-success")
        return ui.TagList()

    @render.plot
    def pct_plot():
        proc = processed_data()
        results = results_obj()
        sel_pct = selected_pct_cols()
        req(proc, results, sel_pct)
        s = settings()
        data = clean_data()
        viz = _make_viz()
        group_col = proc.structure.group_col
        percentages = proc.percentages.loc[data.index].copy()
        display_cols = {col: col.replace('_pct', '') for col in sel_pct}
        display_names = list(display_cols.values())

        if is_twoway():
            fa_col = results.factor_a_col
            fb_col = results.factor_b_col
            pct_plot_data = pd.concat([data[[fa_col, fb_col]].reset_index(drop=True),
                                       percentages[sel_pct].reset_index(drop=True)], axis=1)
            pct_plot_data = pct_plot_data.rename(columns=display_cols)
            tw_stats = {c.replace('_pct', ''): results.twoway_percentages.get(c)
                       for c in sel_pct if c in results.twoway_percentages}
            return viz.plot_twoway_multi_panel(
                data=pct_plot_data, value_cols=display_names,
                factor_a_col=fa_col, factor_b_col=fb_col,
                twoway_results=tw_stats, ncols=3,
                factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
                ylabel='% of Total SL', show_points=s['show_points'],
                plot_type=s['plot_type'])
        else:
            plot_df = pd.concat([data[[group_col]].reset_index(drop=True),
                                 percentages[sel_pct].reset_index(drop=True)], axis=1)
            plot_df_display = plot_df.rename(columns=display_cols)
            stats_dict = {}
            for pct_col in sel_pct:
                dn = pct_col.replace('_pct', '')
                if pct_col in results.percentages_results:
                    stats_dict[dn] = results.percentages_results[pct_col]
            return viz.plot_multi_panel_groups_with_stats(
                plot_df_display, display_names, group_col, stats_dict,
                ncols=3, plot_type=s['plot_type'], log_scale=False,
                show_points=s['show_points'], ylabel='% of Total SL')

    @render.plot
    def pct_interaction_plot():
        req(is_twoway())
        proc = processed_data()
        results = results_obj()
        sel_pct = selected_pct_cols()
        req(proc, results, sel_pct)
        data = clean_data()
        viz = _make_viz()
        percentages = proc.percentages.loc[data.index].copy()
        display_cols = {col: col.replace('_pct', '') for col in sel_pct}
        display_names = list(display_cols.values())
        pct_plot_data = pd.concat([data[[results.factor_a_col, results.factor_b_col]].reset_index(drop=True),
                                   percentages[sel_pct].reset_index(drop=True)], axis=1)
        pct_plot_data = pct_plot_data.rename(columns=display_cols)
        tw_stats = {c.replace('_pct', ''): results.twoway_percentages.get(c)
                   for c in sel_pct if c in results.twoway_percentages}
        return viz.plot_twoway_interaction_multi_panel(
            data=pct_plot_data, value_cols=display_names,
            factor_a_col=results.factor_a_col, factor_b_col=results.factor_b_col,
            twoway_results=tw_stats, ncols=3,
            factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
            ylabel='% of Total SL')

    @render.plot
    def pct_pie_plot():
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        data = clean_data()
        viz = _make_viz()
        group_col = proc.structure.group_col
        percentages = proc.percentages.loc[data.index].copy()
        pct_cols = [c for c in percentages.columns if c.endswith('_pct')]
        all_pct_cols = percentages[pct_cols].mean().nlargest(15).index.tolist()

        if is_twoway():
            comp_group_col = '_factorial_group_'
            comp_data = data.copy()
            comp_data[comp_group_col] = comp_data[results.factor_a_col].astype(str) + ' - ' + comp_data[results.factor_b_col].astype(str)
            pie_df = pd.concat([comp_data[[comp_group_col]].reset_index(drop=True),
                                percentages[all_pct_cols].reset_index(drop=True)], axis=1)
            title_suffix = f" ({results.factor_a_name} x {results.factor_b_name})"
        else:
            comp_group_col = group_col
            pie_df = pd.concat([data[[group_col]].reset_index(drop=True),
                                percentages[all_pct_cols].reset_index(drop=True)], axis=1)
            title_suffix = ""

        return viz.plot_composition_pie_charts(
            pie_df, comp_group_col, all_pct_cols,
            title='Sphingolipid Pool Composition by Group' + title_suffix,
            top_n=10, other_threshold=2.0)

    @render.plot
    def pct_bars_plot():
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        data = clean_data()
        viz = _make_viz()
        group_col = proc.structure.group_col
        percentages = proc.percentages.loc[data.index].copy()
        pct_cols = [c for c in percentages.columns if c.endswith('_pct')]
        all_pct_cols = percentages[pct_cols].mean().nlargest(15).index.tolist()

        if is_twoway():
            comp_group_col = '_factorial_group_'
            comp_data = data.copy()
            comp_data[comp_group_col] = comp_data[results.factor_a_col].astype(str) + ' - ' + comp_data[results.factor_b_col].astype(str)
            pie_df = pd.concat([comp_data[[comp_group_col]].reset_index(drop=True),
                                percentages[all_pct_cols].reset_index(drop=True)], axis=1)
            title_suffix = f" ({results.factor_a_name} x {results.factor_b_name})"
        else:
            comp_group_col = group_col
            pie_df = pd.concat([data[[group_col]].reset_index(drop=True),
                                percentages[all_pct_cols].reset_index(drop=True)], axis=1)
            title_suffix = ""

        return viz.plot_composition_horizontal_bars(
            pie_df, comp_group_col, all_pct_cols,
            title='Sphingolipid Pool Composition Comparison' + title_suffix,
            top_n=15, show_values=True)

    @render.plot
    def pct_stacked_plot():
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        data = clean_data()
        viz = _make_viz()
        group_col = proc.structure.group_col
        percentages = proc.percentages.loc[data.index].copy()
        pct_cols = [c for c in percentages.columns if c.endswith('_pct')]
        all_pct_cols = percentages[pct_cols].mean().nlargest(15).index.tolist()

        if is_twoway():
            comp_group_col = '_factorial_group_'
            comp_data = data.copy()
            comp_data[comp_group_col] = comp_data[results.factor_a_col].astype(str) + ' - ' + comp_data[results.factor_b_col].astype(str)
            pie_df = pd.concat([comp_data[[comp_group_col]].reset_index(drop=True),
                                percentages[all_pct_cols].reset_index(drop=True)], axis=1)
            title_suffix = f" ({results.factor_a_name} x {results.factor_b_name})"
        else:
            comp_group_col = group_col
            pie_df = pd.concat([data[[group_col]].reset_index(drop=True),
                                percentages[all_pct_cols].reset_index(drop=True)], axis=1)
            title_suffix = ""

        return viz.plot_composition_stacked_horizontal(
            pie_df, comp_group_col, all_pct_cols,
            title='Complete Sphingolipid Pool Composition' + title_suffix,
            top_n=12)

    @render.data_frame
    def pct_stats_table():
        proc = processed_data()
        results = results_obj()
        sel_pct = selected_pct_cols()
        req(proc, results, sel_pct)
        if is_twoway():
            from modules.report_generation import get_twoway_differences_summary
            tw_stats = {c: results.twoway_percentages.get(c) for c in sel_pct
                       if c in results.twoway_percentages}
            if tw_stats:
                return render.DataGrid(get_twoway_differences_summary(tw_stats))
            return render.DataGrid(pd.DataFrame())
        else:
            rows = []
            for pct_col in sel_pct:
                sl_name = pct_col.replace('_pct', '')
                res = results.percentages_results.get(pct_col)
                if res:
                    n_sig = 0
                    if res.posthoc_test and res.posthoc_test.pairwise_results is not None:
                        n_sig = res.posthoc_test.pairwise_results['significant'].sum()
                    rows.append({
                        'Sphingolipid': sl_name,
                        'Test': res.main_test.test_type.value,
                        'P-value': f"{res.main_test.pvalue:.4f}",
                        'Significant': 'Yes' if res.main_test.significant else '',
                        'Effect Size': f"{res.main_test.effect_size:.3f}" if res.main_test.effect_size else 'N/A',
                        'Sig. Pairs': n_sig,
                        'APA': format_apa_statistics(res),
                    })
            return render.DataGrid(pd.DataFrame(rows))

    @render.data_frame
    def pct_means_table():
        proc = processed_data()
        sel_pct = selected_pct_cols()
        req(proc, sel_pct)
        data = clean_data()
        group_col = proc.structure.group_col
        percentages = proc.percentages.loc[data.index]
        plot_df = pd.concat([data[[group_col]].reset_index(drop=True),
                             percentages[sel_pct].reset_index(drop=True)], axis=1)
        summary = []
        for pct_col in sel_pct:
            sl = pct_col.replace('_pct', '')
            for group in plot_df[group_col].unique():
                gd = plot_df[plot_df[group_col] == group][pct_col]
                summary.append({'Sphingolipid': sl, 'Group': group,
                                'Mean %': f"{gd.mean():.2f}", 'SD': f"{gd.std():.2f}",
                                'Median %': f"{gd.median():.2f}"})
        if summary:
            df = pd.DataFrame(summary)
            pivot = df.pivot_table(index='Sphingolipid', columns='Group', values='Mean %', aggfunc='first')
            return render.DataGrid(pivot.reset_index())
        return render.DataGrid(pd.DataFrame())

    @render.data_frame
    def pct_raw_table():
        proc = processed_data()
        sel_pct = selected_pct_cols()
        req(proc, sel_pct)
        data = clean_data()
        group_col = proc.structure.group_col
        percentages = proc.percentages.loc[data.index]
        plot_df = pd.concat([data[[group_col]].reset_index(drop=True),
                             percentages[sel_pct].reset_index(drop=True)], axis=1)
        return render.DataGrid(plot_df)

    # =========================================================================
    # Ratios tab
    # =========================================================================

    @render.ui
    def ratio_custom_select():
        proc = processed_data()
        req(proc)
        available = [c for c in proc.ratios.columns if not proc.ratios[c].isna().all()]
        default = available[:6] if len(available) >= 6 else available
        return ui.input_selectize("ratio_custom", "Select ratios",
                                  choices=available, selected=default, multiple=True)

    @reactive.calc
    def selected_ratios():
        proc = processed_data()
        results = results_obj()
        req(proc, results)
        s = settings()
        data = clean_data()
        group_col = proc.structure.group_col
        combined = pd.concat([data[[group_col]], proc.ratios.loc[data.index]], axis=1)
        available = [c for c in proc.ratios.columns if not combined[c].isna().all()]
        quick = input.ratio_quick()

        key_ratios = [
            'C16_to_C24_Cer', 'C24_1_to_C24_0_Cer',
            'Cer_to_DHC_C16', 'Cer_to_DHC_C24', 'SM_to_Cer_C16', 'S1P_to_Sph',
            'very_long_to_long', 'very_long_to_medium', 'long_to_medium',
            'long_to_short', 'medium_to_short',
            'total_Cer_to_SM', 'total_Cer_to_DHC', 'saturated_to_unsaturated',
        ]

        if quick == "All ratios":
            return available
        elif quick == "Significant only":
            if is_twoway():
                sel = [r for r in available if r in results.twoway_ratios
                       and (results.twoway_ratios[r].twoway_result.factor_a_pvalue < s['alpha']
                            or results.twoway_ratios[r].twoway_result.factor_b_pvalue < s['alpha']
                            or results.twoway_ratios[r].twoway_result.interaction_pvalue < s['alpha'])]
            else:
                sel = [r for r in available if r in results.ratios_results
                       and results.ratios_results[r].main_test.significant]
            return sel if sel else available
        elif quick == "Key ratios":
            sel = [r for r in key_ratios if r in available]
            return sel if sel else available[:6]
        else:  # Custom
            custom = input.ratio_custom()
            return list(custom) if custom else []

    @render.ui
    def ratio_sig_info():
        results = results_obj()
        sel = selected_ratios()
        req(results, sel)
        s = settings()
        if is_twoway():
            sig = [r for r in sel if r in results.twoway_ratios
                   and (results.twoway_ratios[r].twoway_result.factor_a_pvalue < s['alpha']
                        or results.twoway_ratios[r].twoway_result.factor_b_pvalue < s['alpha']
                        or results.twoway_ratios[r].twoway_result.interaction_pvalue < s['alpha'])]
        else:
            sig = [r for r in sel if r in results.ratios_results
                   and results.ratios_results[r].main_test.significant]
        if sig:
            return ui.div(ui.strong(f"Significant differences: {', '.join(sig)}"),
                          class_="alert alert-success")
        return ui.TagList()

    @render.plot
    def ratio_plot():
        proc = processed_data()
        results = results_obj()
        sel = selected_ratios()
        req(proc, results, sel)
        s = settings()
        data = clean_data()
        viz = _make_viz()
        group_col = proc.structure.group_col

        if is_twoway():
            ratio_plot_data = pd.concat([data[[results.factor_a_col, results.factor_b_col]].reset_index(drop=True),
                                         proc.ratios[sel].loc[data.index].reset_index(drop=True)], axis=1)
            tw_stats = {r: results.twoway_ratios.get(r) for r in sel if r in results.twoway_ratios}
            return viz.plot_twoway_multi_panel(
                data=ratio_plot_data, value_cols=sel,
                factor_a_col=results.factor_a_col, factor_b_col=results.factor_b_col,
                twoway_results=tw_stats, ncols=3,
                factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
                ylabel='Ratio', show_points=s['show_points'], plot_type=s['plot_type'])
        else:
            combined = pd.concat([data[[group_col]], proc.ratios.loc[data.index]], axis=1)
            stats_dict = {r: results.ratios_results.get(r) for r in sel if r in results.ratios_results}
            return viz.plot_multi_panel_groups_with_stats(
                combined, sel, group_col, stats_dict, ncols=3,
                plot_type=s['plot_type'], log_scale=input.ratio_log(),
                show_points=s['show_points'], ylabel='Ratio')

    @render.plot
    def ratio_interaction_plot():
        req(is_twoway())
        proc = processed_data()
        results = results_obj()
        sel = selected_ratios()
        req(proc, results, sel)
        data = clean_data()
        viz = _make_viz()
        ratio_plot_data = pd.concat([data[[results.factor_a_col, results.factor_b_col]].reset_index(drop=True),
                                     proc.ratios[sel].loc[data.index].reset_index(drop=True)], axis=1)
        tw_stats = {r: results.twoway_ratios.get(r) for r in sel if r in results.twoway_ratios}
        return viz.plot_twoway_interaction_multi_panel(
            data=ratio_plot_data, value_cols=sel,
            factor_a_col=results.factor_a_col, factor_b_col=results.factor_b_col,
            twoway_results=tw_stats, ncols=3,
            factor_a_name=results.factor_a_name, factor_b_name=results.factor_b_name,
            ylabel='Ratio')

    @render.data_frame
    def ratio_stats_table():
        proc = processed_data()
        results = results_obj()
        sel = selected_ratios()
        req(proc, results, sel)
        if is_twoway():
            from modules.report_generation import get_twoway_differences_summary
            tw_stats = {r: results.twoway_ratios.get(r) for r in sel if r in results.twoway_ratios}
            if tw_stats:
                return render.DataGrid(get_twoway_differences_summary(tw_stats))
            return render.DataGrid(pd.DataFrame())
        else:
            rows = []
            for r in sel:
                res = results.ratios_results.get(r)
                if res:
                    n_sig = 0
                    if res.posthoc_test and res.posthoc_test.pairwise_results is not None:
                        n_sig = res.posthoc_test.pairwise_results['significant'].sum()
                    rows.append({
                        'Ratio': r,
                        'Test': res.main_test.test_type.value,
                        'P-value': f"{res.main_test.pvalue:.4f}",
                        'Significant': 'Yes' if res.main_test.significant else '',
                        'Effect Size': f"{res.main_test.effect_size:.3f}" if res.main_test.effect_size else 'N/A',
                        'Sig. Pairs': n_sig,
                        'APA': format_apa_statistics(res),
                    })
            return render.DataGrid(pd.DataFrame(rows))

    @render.data_frame
    def ratio_means_table():
        proc = processed_data()
        sel = selected_ratios()
        req(proc, sel)
        data = clean_data()
        group_col = proc.structure.group_col
        combined = pd.concat([data[[group_col]], proc.ratios.loc[data.index]], axis=1)
        summary = []
        for ratio in sel:
            for group in combined[group_col].unique():
                gd = combined[combined[group_col] == group][ratio]
                summary.append({'Ratio': ratio, 'Group': group,
                                'Mean': f"{gd.mean():.3f}", 'SD': f"{gd.std():.3f}",
                                'Median': f"{gd.median():.3f}"})
        if summary:
            df = pd.DataFrame(summary)
            pivot = df.pivot_table(index='Ratio', columns='Group', values='Mean', aggfunc='first')
            return render.DataGrid(pivot.reset_index())
        return render.DataGrid(pd.DataFrame())

    @render.data_frame
    def ratio_raw_table():
        proc = processed_data()
        sel = selected_ratios()
        req(proc, sel)
        data = clean_data()
        group_col = proc.structure.group_col
        combined = pd.concat([data[[group_col]], proc.ratios.loc[data.index]], axis=1)
        return render.DataGrid(combined[[group_col] + sel])

    # =========================================================================
    # Statistics tab
    # =========================================================================

    @render.ui
    def stats_content():
        results = results_obj()
        proc = processed_data()
        req(results, proc)
        s = settings()
        alpha = s['alpha']

        if results.is_twoway:
            from modules.report_generation import get_twoway_differences_summary
            from modules.statistical_tests import format_twoway_apa

            def _count_sig_tw(results_dict):
                count = 0
                for r in results_dict.values():
                    tw = r.twoway_result
                    if (tw.factor_a_pvalue < alpha or tw.factor_b_pvalue < alpha or tw.interaction_pvalue < alpha):
                        count += 1
                return count

            sig_t = _count_sig_tw(results.twoway_totals)
            sig_s = _count_sig_tw(results.twoway_individual_sl)
            sig_p = _count_sig_tw(results.twoway_percentages)
            sig_r = _count_sig_tw(results.twoway_ratios)

            elements = [
                ui.p(ui.strong(f"Design: Two-way ANOVA ({results.factor_a_name} x {results.factor_b_name})")),
                ui.layout_columns(
                    ui.value_box("Sig. Totals", f"{sig_t}/{len(results.twoway_totals)}"),
                    ui.value_box("Sig. SLs", f"{sig_s}/{len(results.twoway_individual_sl)}"),
                    ui.value_box("Sig. Percentages", f"{sig_p}/{len(results.twoway_percentages)}"),
                    ui.value_box("Sig. Ratios", f"{sig_r}/{len(results.twoway_ratios)}"),
                    col_widths=[3, 3, 3, 3],
                ),
                ui.h4("Totals"),
            ]
            if results.twoway_totals:
                elements.append(ui.output_data_frame("stats_totals_tw_table"))

            elements.append(ui.h4("Individual Sphingolipids (any significant effect)"))
            sig_sl = {k: v for k, v in results.twoway_individual_sl.items()
                      if (v.twoway_result.factor_a_pvalue < alpha or
                          v.twoway_result.factor_b_pvalue < alpha or
                          v.twoway_result.interaction_pvalue < alpha)}
            if sig_sl:
                elements.append(ui.output_data_frame("stats_sl_tw_table"))
                apa_items = []
                for name, result in list(sig_sl.items())[:10]:
                    apa_items.append(ui.p(ui.strong(name), f": {format_twoway_apa(result)}"))
                elements.append(ui.accordion(
                    ui.accordion_panel("APA-Formatted Results", *apa_items),
                    open=False, id="stats_apa_accordion",
                ))
            else:
                elements.append(ui.div("No significant individual sphingolipid differences (two-way ANOVA).",
                                       class_="alert alert-info"))

            elements.append(ui.h4("Ratios"))
            if results.twoway_ratios:
                elements.append(ui.output_data_frame("stats_ratios_tw_table"))

            return ui.TagList(*elements)
        else:
            sig_t = sum(1 for r in results.totals_results.values() if r.main_test.significant)
            sig_s = sum(1 for r in results.individual_sl_results.values() if r.main_test.significant)
            sig_p = sum(1 for r in results.percentages_results.values() if r.main_test.significant)
            sig_r = sum(1 for r in results.ratios_results.values() if r.main_test.significant)

            elements = [
                ui.layout_columns(
                    ui.value_box("Sig. Totals", f"{sig_t}/{len(results.totals_results)}"),
                    ui.value_box("Sig. SLs", f"{sig_s}/{len(results.individual_sl_results)}"),
                    ui.value_box("Sig. Percentages", f"{sig_p}/{len(results.percentages_results)}"),
                    ui.value_box("Sig. Ratios", f"{sig_r}/{len(results.ratios_results)}"),
                    col_widths=[3, 3, 3, 3],
                ),
                ui.h4("Totals"),
                ui.output_data_frame("stats_totals_ow_table"),
                ui.h4("Individual Sphingolipids (Significant)"),
            ]
            sig_sl = {k: v for k, v in results.individual_sl_results.items() if v.main_test.significant}
            if sig_sl:
                elements.append(ui.output_data_frame("stats_sl_ow_table"))
            else:
                elements.append(ui.div("No significant individual sphingolipid differences.",
                                       class_="alert alert-info"))

            elements.append(ui.h4("Percentages (Significant)"))
            sig_pct = {k: v for k, v in results.percentages_results.items() if v.main_test.significant}
            if sig_pct:
                elements.append(ui.output_data_frame("stats_pct_ow_table"))
            else:
                elements.append(ui.div("No significant percentage differences.",
                                       class_="alert alert-info"))

            return ui.TagList(*elements)

    # Two-way stats tables
    @render.data_frame
    def stats_totals_tw_table():
        from modules.report_generation import get_twoway_differences_summary
        results = results_obj()
        req(results, results.is_twoway)
        return render.DataGrid(get_twoway_differences_summary(results.twoway_totals))

    @render.data_frame
    def stats_sl_tw_table():
        from modules.report_generation import get_twoway_differences_summary
        results = results_obj()
        req(results, results.is_twoway)
        alpha = settings()['alpha']
        sig_sl = {k: v for k, v in results.twoway_individual_sl.items()
                  if (v.twoway_result.factor_a_pvalue < alpha or
                      v.twoway_result.factor_b_pvalue < alpha or
                      v.twoway_result.interaction_pvalue < alpha)}
        req(sig_sl)
        return render.DataGrid(get_twoway_differences_summary(sig_sl))

    @render.data_frame
    def stats_ratios_tw_table():
        from modules.report_generation import get_twoway_differences_summary
        results = results_obj()
        req(results, results.is_twoway)
        return render.DataGrid(get_twoway_differences_summary(results.twoway_ratios))

    # One-way stats tables
    @render.data_frame
    def stats_totals_ow_table():
        results = results_obj()
        req(results, not results.is_twoway)
        return render.DataGrid(get_significant_differences_summary(results.totals_results))

    @render.data_frame
    def stats_sl_ow_table():
        results = results_obj()
        req(results, not results.is_twoway)
        sig_sl = {k: v for k, v in results.individual_sl_results.items() if v.main_test.significant}
        req(sig_sl)
        return render.DataGrid(get_significant_differences_summary(sig_sl))

    @render.data_frame
    def stats_pct_ow_table():
        results = results_obj()
        req(results, not results.is_twoway)
        sig_pct = {k: v for k, v in results.percentages_results.items() if v.main_test.significant}
        req(sig_pct)
        sig_display = {k.replace('_pct', ''): v for k, v in sig_pct.items()}
        return render.DataGrid(get_significant_differences_summary(sig_display))

    # =========================================================================
    # Export tab
    # =========================================================================

    @render.download(filename=lambda: "sphingolipid_concentrations.xlsx")
    def download_concentrations():
        proc = processed_data()
        req(proc)
        metadata = _get_metadata(proc)
        conc_df = pd.concat([metadata, proc.concentrations], axis=1)
        buf = create_excel_with_lod_highlighting(conc_df, proc.structure.analyte_lod_rows,
                                                  sheet_name="Concentrations")
        path = os.path.join(tempfile.mkdtemp(), "concentrations.xlsx")
        with open(path, "wb") as f:
            f.write(buf.getvalue())
        return path

    @render.download(filename=lambda: "sphingolipid_totals.csv")
    def download_totals():
        proc = processed_data()
        req(proc)
        metadata = _get_metadata(proc)
        export_df = pd.concat([metadata, proc.totals], axis=1)
        path = os.path.join(tempfile.mkdtemp(), "totals.csv")
        export_df.to_csv(path, index=False)
        return path

    @render.download(filename=lambda: "sphingolipid_ratios.csv")
    def download_ratios():
        proc = processed_data()
        req(proc)
        metadata = _get_metadata(proc)
        export_df = pd.concat([metadata, proc.ratios], axis=1)
        path = os.path.join(tempfile.mkdtemp(), "ratios.csv")
        export_df.to_csv(path, index=False)
        return path

    @render.download(filename=lambda: "sphingolipid_percentages.csv")
    def download_percentages():
        proc = processed_data()
        req(proc)
        metadata = _get_metadata(proc)
        export_df = pd.concat([metadata, proc.percentages], axis=1)
        path = os.path.join(tempfile.mkdtemp(), "percentages.csv")
        export_df.to_csv(path, index=False)
        return path

    @render.download(filename=lambda: f"sphingolipid_all_data_{datetime.now():%Y%m%d}.xlsx")
    def download_full_excel():
        proc = processed_data()
        req(proc)
        metadata = _get_metadata(proc)
        buf = create_full_data_excel_with_highlighting(proc, metadata)
        path = os.path.join(tempfile.mkdtemp(), "all_data.xlsx")
        with open(path, "wb") as f:
            f.write(buf.getvalue())
        return path

    @render.download(filename=lambda: f"statistical_report_{datetime.now():%Y%m%d}.xlsx")
    def download_stats_report():
        rg = report_gen()
        req(rg)
        buf = BytesIO()
        rg.save_excel_report(buf)
        buf.seek(0)
        path = os.path.join(tempfile.mkdtemp(), "report.xlsx")
        with open(path, "wb") as f:
            f.write(buf.getvalue())
        return path

    @render.download(filename=lambda: f"sphingolipid_analysis_{datetime.now():%Y%m%d_%H%M}.zip")
    def download_zip():
        proc = processed_data()
        results = results_obj()
        rg = report_gen()
        req(proc, results, rg)
        s = settings()
        metadata = _get_metadata(proc)
        all_figures = generate_all_export_figures(proc, results, s)
        zip_bytes = create_results_zip(proc, results, all_figures, rg, metadata)
        for fig in all_figures.values():
            if fig:
                plt.close(fig)
        path = os.path.join(tempfile.mkdtemp(), "analysis.zip")
        with open(path, "wb") as f:
            f.write(zip_bytes)
        return path

    @render.plot
    @reactive.event(input.gen_summary)
    def summary_plot():
        proc = processed_data()
        rg = report_gen()
        req(proc, rg)
        s = settings()
        plotter = SignificancePlotter()
        fig = plotter.plot_multi_panel_with_significance(
            proc.sample_data, proc.structure.group_col, rg,
            ['Total_All_Sphingolipids', 'Total_Ceramides', 'Total_Dihydroceramides', 'Total_Sphingomyelins'],
            2, s['plot_type'])
        return fig

    @render.download(filename=lambda: "summary.png")
    def download_summary_png():
        proc = processed_data()
        rg = report_gen()
        req(proc, rg)
        s = settings()
        plotter = SignificancePlotter()
        fig = plotter.plot_multi_panel_with_significance(
            proc.sample_data, proc.structure.group_col, rg,
            ['Total_All_Sphingolipids', 'Total_Ceramides', 'Total_Dihydroceramides', 'Total_Sphingomyelins'],
            2, s['plot_type'])
        path = os.path.join(tempfile.mkdtemp(), "summary.png")
        fig.savefig(path, format='png', dpi=300, bbox_inches='tight')
        plt.close(fig)
        return path

    @render.download(filename=lambda: "summary.pdf")
    def download_summary_pdf():
        proc = processed_data()
        rg = report_gen()
        req(proc, rg)
        s = settings()
        plotter = SignificancePlotter()
        fig = plotter.plot_multi_panel_with_significance(
            proc.sample_data, proc.structure.group_col, rg,
            ['Total_All_Sphingolipids', 'Total_Ceramides', 'Total_Dihydroceramides', 'Total_Sphingomyelins'],
            2, s['plot_type'])
        path = os.path.join(tempfile.mkdtemp(), "summary.pdf")
        fig.savefig(path, format='pdf', dpi=300, bbox_inches='tight')
        plt.close(fig)
        return path


# =============================================================================
# App
# =============================================================================

app = App(app_ui, server)
